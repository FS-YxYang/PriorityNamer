from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import requests

now = datetime.now()
thisMonth = now.month
nextMonth =(now + timedelta(days=31)).month
timestamp = now.strftime("%Y-%m-%d %H:%M")
testtimestamp = now.strftime("%Y-%m-%d %H")
print(f"{nextMonth}月")

'''
sheet_id = "17yH30whA5VllYxam5LfTArEAq7sbF87cxDmPAZmm4SY"
url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

response = requests.get(url)
with open(f"test_sheet_{testtimestamp}.xlsx", "wb") as f:
    f.write(response.content)
'''

SC = load_workbook(f"test_sheet_{testtimestamp}.xlsx")
ThisMonthSheet = SC[f'{thisMonth}月配信分のスケジュール']

#For columns in the sheet, find how many cells with the HEX color #feff01 and make one dictionary for each column that has more than one cell with this color
column_dicts = {}

# Get the maximum column and row
max_col = ThisMonthSheet.max_column
max_row = ThisMonthSheet.max_row

target_rgb = ThisMonthSheet.cell(row=1, column=1).fill.start_color.rgb

# Iterate through each column
for col in range(2, max_col):
    #print(ThisMonthSheet.cell(row=2, column=col).value)
    yellow_cells = []
    
    # Check each cell in the column
    for row in range(1, max_row + 1):
        cell = ThisMonthSheet.cell(row=row, column=col)
        cellColor = cell.fill.start_color.rgb
        if cellColor == target_rgb:
            yellow_cells.append((cell.coordinate, cell.value))
    
    # If we found more than one yellow cell in this column, create a dictionary
    if len(yellow_cells) >= 1:
        column_dicts[col] = {
            'yellow_cells': yellow_cells,
            'count': len(yellow_cells),
            'column_name': ThisMonthSheet.cell(row=2, column=col).value
        }

# Print the results
for col, data in column_dicts.items():
    print(f"Column {col} has {data['count']} yellow cells with column name: {data['column_name']}")